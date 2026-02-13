#!/usr/bin/env python3
"""
Query Channels DVR API for recordings and display/export to Excel.

Usage:
    python query_channels_recordings.py
    python query_channels_recordings.py -columns title,created_at,path
    python query_channels_recordings.py -excel -file recordings.xlsx
    python query_channels_recordings.py -excel \\
        -file output.xlsx -columns title,path,completed
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import requests  # noqa: E402
from py_captions_for_channels.config import CHANNELS_API_URL  # noqa: E402


def get_channels_api_url():
    """Get Channels API URL from config."""
    return CHANNELS_API_URL.rstrip("/")


def fetch_recordings(api_url):
    """Fetch all recordings from Channels DVR API."""
    try:
        response = requests.get(f"{api_url}/dvr/files", timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching recordings: {e}", file=sys.stderr)
        sys.exit(1)


# Map common lowercase names to actual API field names
FIELD_MAPPINGS = {
    "path": "Path",
    "created_at": "CreatedAt",
    "updated_at": "UpdatedAt",
    "duration": "Duration",
    "id": "ID",
    "job_id": "JobID",
    "rule_id": "RuleID",
    "group_id": "GroupID",
    "channel_number": "ChannelNumber",
    "device_id": "DeviceID",
    "version": "Version",
    "job_time": "JobTime",
    "job_duration": "JobDuration",
    "title": "Airing.Title",
    "summary": "Airing.Summary",
    "aired_at": "Airing.Time",
    "original_date": "Airing.OriginalDate",
    "channel": "Airing.Channel",
    "series_id": "Airing.SeriesID",
    "program_id": "Airing.ProgramID",
    "episode_number": "Airing.EpisodeNumber",
    "release_year": "Airing.ReleaseYear",
    "image": "Airing.Image",
    "categories": "Airing.Categories",
    "genres": "Airing.Genres",
}


def get_nested_value(obj, path):
    """Get a nested value from a dict using dot notation."""
    keys = path.split(".")
    value = obj
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
            if value is None:
                return None
        else:
            return None
    return value


def get_recording_value(recording, column):
    """Get a value from recording, handling field name mapping."""
    # Map common lowercase names to actual API fields
    field_name = FIELD_MAPPINGS.get(column.lower(), column)

    # Handle nested fields (e.g., "Airing.Title")
    if "." in field_name:
        return get_nested_value(recording, field_name)

    # Direct field access
    return recording.get(field_name)


def format_value(value):
    """Format a value for display."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return str(value)


def format_datetime(timestamp):
    """Format a Unix timestamp to readable datetime."""
    if timestamp is None or timestamp == "":
        return ""
    try:
        if isinstance(timestamp, str):
            timestamp = int(timestamp)
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return str(timestamp)


def display_text_output(recordings, columns):
    """Display recordings as text output."""
    if not recordings:
        print("No recordings found.")
        return

    # Print header
    print("\t".join(columns))
    print("-" * 80)

    # Print each recording
    for recording in recordings:
        values = []
        for col in columns:
            value = get_recording_value(recording, col)

            # Format datetime fields
            if col.lower() in ["created_at", "updated_at", "aired_at"]:
                value = format_datetime(value)
            else:
                value = format_value(value)

            values.append(value)

        print("\t".join(values))


def export_to_excel(recordings, columns, filename):
    """Export recordings to Excel file."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("Error: openpyxl is required for Excel export.", file=sys.stderr)
        print("Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not recordings:
        print("No recordings to export.")
        return

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recordings"

    # Write header row with formatting
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_idx, recording in enumerate(recordings, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = get_recording_value(recording, col_name)

            # Format datetime fields
            if col_name.lower() in ["created_at", "updated_at", "aired_at"]:
                value = format_datetime(value)
            else:
                value = format_value(value)

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, start=1):
        max_length = len(col_name)
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save workbook
    wb.save(filename)
    print(f"Exported {len(recordings)} recordings to {filename}")


def main():
    parser = argparse.ArgumentParser(
        description="Query Channels DVR API for recordings",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s
  %(prog)s -columns title,created_at,path
  %(prog)s -excel -file recordings.xlsx
  %(prog)s -excel -file output.xlsx -columns title,path,completed
        """,
    )

    parser.add_argument(
        "-columns",
        type=str,
        help=(
            "Comma-delimited list of column names "
            "(default: title,created_at,duration,path)"
        ),
    )

    parser.add_argument(
        "-excel",
        action="store_true",
        help="Export to Excel format (requires -file argument)",
    )

    parser.add_argument(
        "-file",
        type=str,
        help="Output filename for Excel export (required when using -excel)",
    )

    args = parser.parse_args()

    # Validate arguments
    if args.excel and not args.file:
        parser.error("-excel requires -file argument")

    # Determine columns to display
    default_columns = ["title", "created_at", "duration", "path"]

    if args.columns:
        columns = [col.strip() for col in args.columns.split(",")]
    else:
        columns = default_columns

    # Fetch recordings
    api_url = get_channels_api_url()
    print(f"Fetching recordings from {api_url}...", file=sys.stderr)
    recordings = fetch_recordings(api_url)
    print(f"Found {len(recordings)} recordings.", file=sys.stderr)

    # Output results
    if args.excel:
        export_to_excel(recordings, columns, args.file)
    else:
        display_text_output(recordings, columns)


if __name__ == "__main__":
    main()
